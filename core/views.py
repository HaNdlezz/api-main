from django.shortcuts import render_to_response, render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.core.urlresolvers import reverse, reverse_lazy
from core.models import *
import datetime
import requests
from itertools import chain
import json
import os
from unidecode import unidecode
import base64
from openpyxl import load_workbook
# from Crypto.Cipher import AES
#import requests
import ast
import random 
from django.views.decorators.csrf import csrf_exempt
import pdb


def login_user(request):
    template_name = 'login.html'
    logout(request)
    data = {}
    username = password = ''
    if request.POST:
        username = request.POST['Username']
        password = request.POST['Password']
        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                return HttpResponseRedirect(reverse_lazy('index'))
#            else :
#                messages.warning(request, ('Usuario inactivo!'))
#        else :
#            messages.error(request, ('Usuario no existe!'))
        data["no_login"] = "ok"
    return render(request, template_name, data)

@login_required(login_url='/')
def index(request):
    data = {}
    template_name = 'index.html'
    data["quantity_students"] = "0"
    data["alumnos"] = alumno.objects.all()
    return render(request, template_name, data)

@login_required(login_url='/')
def upload(request):
    template_name = 'index.html'
    data = {}
    excel = request.FILES['excel']
    fs_excel = FileSystemStorage()
    filename_excel = fs_excel.save(excel.name.lower(),excel)
    excel_url = fs_excel.url(filename_excel)
    excel_url = excel_url[1:].capitalize()
    wb = load_workbook(excel_url)
    quantity_students = 0
    for x in wb.get_sheet_names():
        for y in wb[x].iter_rows():
            row = []
            for z in y:
                try:
                    row.append(unidecode(z.value))
                except:
                    row.append(str(z.value))
            if len(row) < 29:
                data["quantity_students"] = "-1"
                os.system('rm Media/*.xlsx')
                return render(request, template_name, data)
            if len(row) > 29:
                data["quantity_students"] = "-2"
                os.system('rm Media/*.xlsx')
                return render(request, template_name, data)
            if len(row) == 29 and quantity_students > 0:
                alumno1 = alumno.objects.create(rut=row[1],
                                                nombre=row[2],
                                                fecha_nacimiento=row[3][:10],
                                                campus=row[4],
                                                unidad_academica=row[5],
                                                programa=row[6],
                                                carrera=row[7],
                                                especialidad=row[8],
                                                mencion=row[9],
                                                estado=row[10],
                                                matriculado=row[11],
                                                bloqueo_sin_matricula=row[12],
                                                retencion_financiera=row[13],
                                                periodo_ultima_actualizacion=row[14],
                                                periodo_ctgl=row[15],
                                                periodo_admision=row[16],
                                                tipo_alumno=row[17],
                                                ppa=row[18],
                                                pga_programa=row[19],
                                                pga_historico=row[20],
                                                asignaturas_aprobadas=row[21],
                                                creditos_aprobados_programa=row[22],
                                                via_titulacion=row[23],
                                                domicilio=row[24],
                                                fono_celular=row[25],
                                                fono_particular=row[26],
                                                correo_unab=row[27],
                                                correo_personal=row[28])
                alumno1.save()
            quantity_students+=1
    os.system('rm *.xlsx')
    data["quantity_students"] = str(quantity_students)
    return render(request, template_name, data)

@login_required(login_url='/')
def carga_evaluacion_docente(request):
    data = {}
    template_name = 'carga_evaluacion_docente.html'
    data["quantity_teachers"] = "0"
    return render(request, template_name, data)

@login_required(login_url='/')
def upload_docente(request):
    template_name = 'carga_evaluacion_docente.html'
    quantity_teachers = 0
    data = {}
    print request.FILES
    excel = request.FILES['excel']
    fs_excel = FileSystemStorage()
    filename_excel = fs_excel.save(excel.name.lower(),excel)
    excel_url = fs_excel.url(filename_excel)
    excel_url = excel_url[1:].capitalize()
    wb = load_workbook(excel_url)
    for x in wb.get_sheet_names():
        for y in wb[x].iter_rows():
            row = []
            for z in y:
                print z.value, type(z.value)
                try:
                    row.append(unidecode(z.value))
                except:
                    row.append(str(z.value))
            print row
            if len(docente.objects.filter(ano=str(row[8]))) > 0 and quantity_teachers == 0:
                data["error"] = "1"
                os.system('rm Media/*.xlsx')
                return render(request, template_name, data)
            if len(row) < 141:
                data["quantity_teachers"] = "-1"
                os.system('rm Media/*.xlsx')
                return render(request, template_name, data)
            if len(row) > 141:
                data["quantity_teachers"] = "-2"
                os.system('rm Media/*.xlsx')
                return render(request, template_name, data)
            if len(row) == 141 and quantity_teachers > 0:
                docente1 = docente.objects.create(nombre=row[0],
                                                mail=row[1],
                                                run=row[2][:10],
                                                titulo=row[3],
                                                grado=row[4],
                                                jornada=row[5],
                                                sede=row[6],
                                                asignatura=row[7],
                                                ano=row[8],
                                                semestre=row[9],
                                                opa1=row[10],
                                                opa2=row[11],
                                                opa3=row[12],
                                                opa4=row[13],
                                                opa5=row[14],
                                                opb1=row[15],
                                                opb2=row[16],
                                                opb3=row[17],
                                                opb4=row[18],
                                                opb5=row[19],
                                                opc1=row[20],
                                                opc2=row[21],
                                                opc3=row[22],
                                                opc4=row[23],
                                                opc5=row[24],
                                                opd1=row[25],
                                                opd2=row[26],
                                                opd3=row[27],
                                                opd4=row[28],
                                                opd5=row[29],
                                                ope1=row[30],
                                                ope2=row[31],
                                                ope3=row[32],
                                                ope4=row[33],
                                                ope5=row[34],
                                                opf1=row[35],
                                                opf2=row[36],
                                                opf3=row[37],
                                                opf4=row[38],
                                                opf5=row[39],
                                                opg1=row[40],
                                                opg2=row[41],
                                                opg3=row[42],
                                                opg4=row[43],
                                                opg5=row[44],
                                                oph1=row[45],
                                                oph2=row[46],
                                                oph3=row[47],
                                                oph4=row[48],
                                                oph5=row[49],
                                                opi1=row[50],
                                                opi2=row[51],
                                                opi3=row[52],
                                                opi4=row[53],
                                                opi5=row[54],
                                                opj1=row[55],
                                                opj2=row[56],
                                                opj3=row[57],
                                                opj4=row[58],
                                                opj5=row[59],
                                                opk1=row[60],
                                                opk2=row[61],
                                                opk3=row[62],
                                                opk4=row[63],
                                                opk5=row[64],
                                                va1=row[65],
                                                va2=row[66],
                                                va3=row[67],
                                                va4=row[68],
                                                va5=row[69],
                                                vpb1=row[70],
                                                vpb2=row[71],
                                                vpb3=row[72],
                                                vpb4=row[73],
                                                vpb5=row[74],
                                                vpc1=row[75],
                                                vpc2=row[76],
                                                vpc3=row[77],
                                                vpc4=row[78],
                                                vpc5=row[79],
                                                vpd1=row[80],
                                                vpd2=row[81],
                                                vpd3=row[82],
                                                vpd4=row[83],
                                                vpd5=row[84],
                                                ea1=row[85],
                                                ea2=row[86],
                                                ea3=row[87],
                                                ea4=row[88],
                                                ea5=row[89],
                                                eb1=row[90],
                                                eb2=row[91],
                                                eb3=row[92],
                                                eb4=row[93],
                                                eb5=row[94],
                                                ec1=row[95],
                                                ec2=row[96],
                                                ec3=row[97],
                                                ec4=row[98],
                                                ec5=row[99],
                                                ed1=row[100],
                                                ed2=row[101],
                                                ed3=row[102],
                                                ed4=row[103],
                                                ed5=row[104],
                                                ee1=row[105],
                                                ee2=row[106],
                                                ee3=row[107],
                                                ee4=row[108],
                                                ee5=row[109],
                                                ef1=row[110],
                                                ef2=row[111],
                                                ef3=row[112],
                                                ef4=row[113],
                                                ef5=row[114],
                                                eg1=row[115],
                                                eg2=row[116],
                                                eg3=row[117],
                                                eg4=row[118],
                                                eg5=row[119],
                                                opn1=row[120],
                                                opn2=row[121],
                                                opn3=row[122],
                                                opn4=row[123],
                                                opn5=row[124],
                                                vn1=row[125],
                                                vn2=row[126],
                                                vn3=row[127],
                                                vn4=row[128],
                                                vn5=row[129],
                                                es1=row[130],
                                                es2=row[131],
                                                es3=row[132],
                                                es4=row[133],
                                                es5=row[134],
                                                nivel1=row[135],
                                                nivel2=row[136],
                                                nivel3=row[137],
                                                nivel4=row[138],
                                                nivel5=row[139],
                                                evaluacion=row[140])
                                                
                docente1.save()
            quantity_teachers+=1
    data["quantity_teachers"] = str(quantity_teachers-1)
    return render(request, template_name, data)

def search(request):
    response = {"teachers":[]}
    teachers = docente.objects.all()
    for teacher in teachers:
        aux = []
        aux.append(teacher.jornada)
        aux.append(teacher.nombre)
        aux.append(teacher.asignatura)
        aux.append(teacher.sede)
        aux.append(teacher.semestre)
        aux.append(teacher.ano)
        aux.append(teacher.pk)
        response["teachers"].append(aux)
    return HttpResponse(
            json.dumps(response),
            content_type="application/json"
            )

@csrf_exempt
def details_evaluacion_docente(request):
    response = {}
    pkRecibida = int(request.POST["pk"])
    docenteEvaluation = docente.objects.get(pk=pkRecibida)
    response["anio"] = docenteEvaluation.ano
    response["semestre"] = docenteEvaluation.semestre
    response["nombreCurso"] = docenteEvaluation.asignatura
    response["jornada"] = docenteEvaluation.jornada
    response["sede"] = docenteEvaluation.sede
    response["nombreProfesor"] = docenteEvaluation.nombre
    response["rut"] = docenteEvaluation.run
    response["email"] = docenteEvaluation.mail
    response["titulo"] = docenteEvaluation.titulo
    response["gradoAcademico"] = docenteEvaluation.grado
    response["opn1"] = docenteEvaluation.opn1
    response["opn2"] = docenteEvaluation.opn2
    response["opn3"] = docenteEvaluation.opn3
    response["opn4"] = docenteEvaluation.opn4
    response["opn5"] = docenteEvaluation.opn5
    response["vn1"] = docenteEvaluation.vn1
    response["vn2"] = docenteEvaluation.vn2
    response["vn3"] = docenteEvaluation.vn3
    response["vn4"] = docenteEvaluation.vn4
    response["vn5"] = docenteEvaluation.vn5
    response["es1"] = docenteEvaluation.es1
    response["es2"] = docenteEvaluation.es2
    response["es3"] = docenteEvaluation.es3
    response["es4"] = docenteEvaluation.es4
    response["es5"] = docenteEvaluation.es5
    response["nivel1"] = docenteEvaluation.nivel1
    response["nivel2"] = docenteEvaluation.nivel2
    response["nivel3"] = docenteEvaluation.nivel3
    response["nivel4"] = docenteEvaluation.nivel4
    response["nivel5"] = docenteEvaluation.nivel5
    response["evaluacion"] = docenteEvaluation.evaluacion
    return HttpResponse(
        json.dumps(response),
        content_type="application/json"
        )

@csrf_exempt
def API_ANUAL(request):
	response = {}
	variableRecibida = request.POST["anio"]
	primerSemestreRegular = variableRecibida + "10"
	segundoSemestreRegular = variableRecibida + "20"
	primerTrimestreEspecial = variableRecibida + "05"
	segundoTrimestreEspecial = variableRecibida + "15"
	tercerTrimestreEspecial = variableRecibida + "25"

	cantidadAlumnosDiurnos = len(alumno.objects.filter(programa="UNAB11500", periodo_admision=segundoSemestreRegular))+len(alumno.objects.filter(programa="UNAB11500", periodo_admision=primerSemestreRegular))
	cantidadAlumnosVespertinos = len(alumno.objects.filter(programa="UNAB21500", periodo_admision=segundoSemestreRegular))+len(alumno.objects.filter(programa="UNAB21500", periodo_admision=primerSemestreRegular))
	cantidadAlumnosAdvance = len(alumno.objects.filter(programa="UNAB21503", periodo_admision=tercerTrimestreEspecial))+len(alumno.objects.filter(programa="UNAB21503", periodo_admision=segundoTrimestreEspecial))+len(alumno.objects.filter(programa="UNAB21503", periodo_admision=primerTrimestreEspecial))
	
	dataTable = []
	aux=0
	aux2=0
	aux3=0
	if str(request.POST["verificar"])==str(0):
		for i in range (1993,2019):
			aux = str(i)+str(10)
			aux2 = str(i)+str(20)
			count_diu = (len(alumno.objects.filter(programa="UNAB11500", periodo_admision=aux))+len(alumno.objects.filter(programa="UNAB11500", periodo_admision=aux2)))
			count_vesp = (len(alumno.objects.filter(programa="UNAB21500", periodo_admision=aux))+len(alumno.objects.filter(programa="UNAB21500", periodo_admision=aux2)))
			aux = str(i)+"05"
			aux2 = str(i)+"15"
			aux3 = str(i)+"25"
			count_adv = (len(alumno.objects.filter(programa="UNAB21503", periodo_admision=aux))+len(alumno.objects.filter(programa="UNAB21503", periodo_admision=aux2))+len(alumno.objects.filter(programa="UNAB21503", periodo_admision=aux3)))
			dataTable.append([i, count_diu, count_vesp, count_adv])

		dataTable=dataTable[::-1]
		response["dataTable"] = dataTable
	response["diurnos"] = cantidadAlumnosDiurnos
	response["vespertinos"] = cantidadAlumnosVespertinos
	response["advance"] = cantidadAlumnosAdvance
	return HttpResponse(
			json.dumps(response),
			content_type="application/json"
			)

@csrf_exempt
def API_GENERAL(request):
	response = {}
	#Consulta de alumnos diurnos
	activosD = alumno.objects.filter(programa="UNAB11500", estado="ACTIVO")
	response["activosD"] = len(activosD)
	intercambiosD = alumno.objects.filter(programa="UNAB11500", estado="ALUMNO DE INTERCAMBIO")
	response["intercambiosD"] = len(intercambiosD)
	bloqueadosD = alumno.objects.filter(programa="UNAB11500", estado="BLOQUEADO ACADEMICAMENTE")
	response["bloqueadosD"] = len(bloqueadosD)
	desertoresD = alumno.objects.filter(programa="UNAB11500", estado="DESERTOR")
	response["desertoresD"] = len(desertoresD)
	egresadosD = alumno.objects.filter(programa="UNAB11500", estado="EGRESO")
	response["egresadosD"] = len(egresadosD)
	eliminadosD = alumno.objects.filter(programa="UNAB11500", estado="ELIMINADO ACADEMICAMENTE")
	response["eliminadosD"] = len(eliminadosD)
	finalizadosD = alumno.objects.filter(programa="UNAB11500", estado="FINALIZADO")
	response["finalizadosD"] = len(finalizadosD)
	renunciadosD = alumno.objects.filter(programa="UNAB11500", estado="RENUNCIA VACANTE")
	response["renunciadosD"] = len(renunciadosD)
	requisitosAcademicosFinalizadosD = alumno.objects.filter(programa="UNAB11500", estado="REQUISITOS ACADEM. FINALIZADOS")
	response["requisitosAcademicosFinalizadosD"] = len(requisitosAcademicosFinalizadosD)
	resciliacionD = alumno.objects.filter(programa="UNAB11500", estado="RESCILIACION")
	response["resciliacionD"] = len(resciliacionD)
	rDefinitivoD = alumno.objects.filter(programa="UNAB11500", estado="RETIRO DEFINITIVO")
	response["rDefinitivoD"] = len(rDefinitivoD)
	rTemporalD = alumno.objects.filter(programa="UNAB11500", estado="RETIRO TEMPORAL")
	response["rTemporalD"] = len(rTemporalD)
	retractadosD = alumno.objects.filter(programa="UNAB11500", estado="RETRACTO")
	response["retractadosD"] = len(retractadosD)
	#Consulta de alumnos vespertinos
	activosV = alumno.objects.filter(programa="UNAB21500", estado="ACTIVO")
	response["activosV"] = len(activosV)
	intercambiosV = alumno.objects.filter(programa="UNAB21500", estado="ALUMNO DE INTERCAMBIO")
	response["intercambiosV"] = len(intercambiosV)
	bloqueadosV = alumno.objects.filter(programa="UNAB21500", estado="BLOQUEADO ACADEMICAMENTE")
	response["bloqueadosV"] = len(bloqueadosV)
	desertoresV = alumno.objects.filter(programa="UNAB21500", estado="DESERTOR")
	response["desertoresV"] = len(desertoresV)
	egresadosV = alumno.objects.filter(programa="UNAB21500", estado="EGRESO")
	response["egresadosV"] = len(egresadosV)
	eliminadosV = alumno.objects.filter(programa="UNAB21500", estado="ELIMINADO ACADEMICAMENTE")
	response["eliminadosV"] = len(eliminadosV)
	finalizadosV = alumno.objects.filter(programa="UNAB21500", estado="FINALIZADO")
	response["finalizadosV"] = len(finalizadosV)
	renunciadosV = alumno.objects.filter(programa="UNAB21500", estado="RENUNCIA VACANTE")
	response["renunciadosV"] = len(renunciadosV)
	requisitosAcademicosFinalizadosV = alumno.objects.filter(programa="UNAB21500", estado="REQUISITOS ACADEM. FINALIZADOS")
	response["requisitosAcademicosFinalizadosV"] = len(requisitosAcademicosFinalizadosV)
	resciliacionV = alumno.objects.filter(programa="UNAB21500", estado="RESCILIACION")
	response["resciliacionV"] = len(resciliacionV)
	rDefinitivoV = alumno.objects.filter(programa="UNAB21500", estado="RETIRO DEFINITIVO")
	response["rDefinitivoV"] = len(rDefinitivoV)
	rTemporalV = alumno.objects.filter(programa="UNAB21500", estado="RETIRO TEMPORAL")
	response["rTemporalV"] = len(rTemporalV)
	retractadosV = alumno.objects.filter(programa="UNAB21500", estado="RETRACTO")
	response["retractadosV"] = len(retractadosV)
	#Consulta de alumnos advance
	activosA = alumno.objects.filter(programa="UNAB21503", estado="ACTIVO")
	response["activosA"] = len(activosA)
	intercambiosA = alumno.objects.filter(programa="UNAB21503", estado="ALUMNO DE INTERCAMBIO")
	response["intercambiosA"] = len(intercambiosA)
	bloqueadosA = alumno.objects.filter(programa="UNAB21503", estado="BLOQUEADO ACADEMICAMENTE")
	response["bloqueadosA"] = len(bloqueadosA)
	desertoresA = alumno.objects.filter(programa="UNAB21503", estado="DESERTOR")
	response["desertoresA"] = len(desertoresA)
	egresadosA = alumno.objects.filter(programa="UNAB21503", estado="EGRESO")
	response["egresadosA"] = len(egresadosA)
	eliminadosA = alumno.objects.filter(programa="UNAB21503", estado="ELIMINADO ACADEMICAMENTE")
	response["eliminadosA"] = len(eliminadosA)
	finalizadosA = alumno.objects.filter(programa="UNAB21503", estado="FINALIZADO")
	response["finalizadosA"] = len(finalizadosA)
	renunciadosA = alumno.objects.filter(programa="UNAB21503", estado="RENUNCIA VACANTE")
	response["renunciadosA"] = len(renunciadosA)
	requisitosAcademicosFinalizadosA = alumno.objects.filter(programa="UNAB21503", estado="REQUISITOS ACADEM. FINALIZADOS")
	response["requisitosAcademicosFinalizadosA"] = len(requisitosAcademicosFinalizadosA)
	resciliacionA = alumno.objects.filter(programa="UNAB21503", estado="RESCILIACION")
	response["resciliacionA"] = len(resciliacionA)
	rDefinitivoA = alumno.objects.filter(programa="UNAB21503", estado="RETIRO DEFINITIVO")
	response["rDefinitivoA"] = len(rDefinitivoA)
	rTemporalA = alumno.objects.filter(programa="UNAB21503", estado="RETIRO TEMPORAL")
	response["rTemporalA"] = len(rTemporalA)
	retractadosA = alumno.objects.filter(programa="UNAB21503", estado="RETRACTO")
	response["retractadosA"] = len(retractadosA)

	print response
	return HttpResponse(
			json.dumps(response),
			content_type="application/json"
			)

@csrf_exempt
def API_POST(request):
	response = {}
	print request.POST
	if request.POST["hola"] == "1234":
		response["ok"] = "ok"
	return HttpResponse(
			json.dumps(response),
			content_type="application/json"
			)